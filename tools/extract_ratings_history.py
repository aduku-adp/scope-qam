#!/usr/bin/env python3
"""Extract rating assessments from MASTER sheet to JSON and load into Postgres."""

from __future__ import annotations

import json
import os
import re
import hashlib
import psycopg2
from dataclasses import dataclass
from pathlib import Path
from typing import Any
from zipfile import BadZipFile

from openpyxl import load_workbook

TARGET_SHEET_NAME = "MASTER"
DATA_DIR = Path("../data")
EXCEL_EXTENSIONS = (".xlsm", ".xlsx", ".xltm", ".xltx")
ENV_FILE = Path(".env")
DATA_ONLY = True


@dataclass
class DbConfig:
    host: str
    port: int
    user: str
    password: str
    dbname: str


def parse_env_file(env_path: Path) -> dict[str, str]:
    values: dict[str, str] = {}
    if not env_path.exists():
        return values

    for raw_line in env_path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        values[key.strip()] = value.strip().strip('"').strip("'")
    return values


def load_db_config(env_path: Path) -> DbConfig:
    file_values = parse_env_file(env_path)

    def get_value(*keys: str, default: str | None = None) -> str:
        for key in keys:
            if os.getenv(key):
                return os.getenv(key, "")
            if key in file_values:
                return file_values[key]
        if default is None:
            raise ValueError(
                f"Missing required DB config. Checked keys: {', '.join(keys)}"
            )
        return default

    return DbConfig(
        host=get_value("PG_HOST", default="localhost"),
        port=int(get_value("PG_PORT", default="5432")),
        user=get_value("PG_USER", default="postgres"),
        password=get_value("PG_PASSWORD", default="postgres"),
        dbname=get_value("DB_NAME", "PG_DATABASE", default="qam_db"),
    )


def parse_float_value(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, str):
        text = value.strip()
        if not text or text.lower() in {"no data", "n/a", "na"}:
            return None
        try:
            return round(float(text.replace(",", "")), 2)
        except ValueError:
            return None
    if isinstance(value, (int, float)):
        return round(float(value), 2)
    return None


def parse_notches(value: Any) -> int | None:
    if value is None:
        return None
    match = re.search(r"-?\d+", str(value))
    if not match:
        return None
    return int(match.group(0))


def normalize_metric_name(metric_name: str) -> str:
    normalized = metric_name.strip().lower()
    normalized = re.sub(r"[^a-z0-9]+", "_", normalized)
    return normalized.strip("_")


def get_label_value(master_sheet, label: str) -> Any:
    for row_idx in range(1, master_sheet.max_row + 1):
        if master_sheet.cell(row=row_idx, column=2).value == label:
            return master_sheet.cell(row=row_idx, column=3).value
    return None


def extract_credit_metrics(master_sheet) -> list[dict[str, Any]]:
    header_row = None
    for row_idx in range(1, master_sheet.max_row + 1):
        if master_sheet.cell(row=row_idx, column=2).value == "[Scope Credit Metrics]":
            header_row = row_idx
            break
    if header_row is None:
        return []

    years: list[str] = []
    col_idx = 3
    while True:
        year_value = master_sheet.cell(row=header_row, column=col_idx).value
        if year_value in (None, ""):
            break
        years.append(str(year_value))
        col_idx += 1

    lock_col = 3 + len(years)
    metrics: list[dict[str, Any]] = []

    row_idx = header_row + 1
    while row_idx <= master_sheet.max_row:
        metric_label = master_sheet.cell(row=row_idx, column=2).value
        if metric_label in (None, ""):
            break

        values = []
        for offset, year in enumerate(years):
            raw_value = master_sheet.cell(row=row_idx, column=3 + offset).value
            values.append({"year": year, "value": parse_float_value(raw_value)})

        lock_value = master_sheet.cell(row=row_idx, column=lock_col).value
        locked = str(lock_value).strip().lower() == "locked"

        metrics.append(
            {
                "metric": normalize_metric_name(str(metric_label)),
                "values": values,
                "locked": locked,
            }
        )
        row_idx += 1

    return metrics


def extract_workbook(workbook_path: Path, data_only: bool) -> dict[str, Any]:
    workbook = load_workbook(workbook_path, data_only=data_only)
    if TARGET_SHEET_NAME not in workbook.sheetnames:
        raise ValueError(f"Sheet '{TARGET_SHEET_NAME}' not found in workbook")

    ws = workbook[TARGET_SHEET_NAME]

    methodologies = []
    for col_idx in range(3, ws.max_column + 1):
        value = ws.cell(row=5, column=col_idx).value
        if value in (None, ""):
            continue
        methodologies.append(str(value))

    # The source workbook stores the label "CorporateSector" in column B and the
    # industry value in column C on row 3. The expected JSON keeps both fields.
    corporate_sector_label = ws.cell(row=3, column=2).value
    industry_value = ws.cell(row=3, column=3).value

    return {
        "entity_information": {
            "name": get_label_value(ws, "Rated entity"),
            "corporate_sector": corporate_sector_label,
            "industry": industry_value,
            "country_of_origin": get_label_value(ws, "Country of origin"),
            "reporting_currency": get_label_value(ws, "Reporting Currency/Units"),
            "accounting_principles": get_label_value(ws, "Accounting principles"),
            "fiscal_year_end": get_label_value(ws, "End of business year"),
        },
        "methodology": {
            "rating_methodologies_applied": methodologies,
        },
        "industry_risk": {
            "industry_classification": get_label_value(ws, "Industry risk"),
            "industry_risk_score": get_label_value(ws, "Industry risk score"),
            "industry_weight": parse_float_value(
                get_label_value(ws, "Industry weight")
            ),
            "segmentation_criteria": get_label_value(ws, "Segmentation criteria"),
        },
        "business_risk_profile": {
            "overall_score": get_label_value(ws, "Business risk profile"),
            "components": {
                "blended_industry_risk_profile": get_label_value(
                    ws, "(Blended) Industry risk profile"
                ),
                "competitive_positioning": get_label_value(
                    ws, "Competitive Positioning"
                ),
                "market_share": get_label_value(ws, "Market share"),
                "diversification": get_label_value(ws, "Diversification"),
                "operating_profitability": get_label_value(
                    ws, "Operating profitability"
                ),
                "sector_company_specific_factors_1": get_label_value(
                    ws, "Sector/company-specific factors (1)"
                ),
                "sector_company_specific_factors_2": get_label_value(
                    ws, "Sector/company-specific factors (2)"
                ),
            },
        },
        "financial_risk_profile": {
            "overall_score": get_label_value(ws, "Financial risk profile"),
            "components": {
                "leverage": get_label_value(ws, "Leverage"),
                "interest_cover": get_label_value(ws, "Interest cover"),
                "cash_flow_cover": get_label_value(ws, "Cash flow cover"),
                "liquidity_adjustment_notches": parse_notches(
                    get_label_value(ws, "Liquidity")
                ),
            },
        },
        "credit_metrics": extract_credit_metrics(ws),
    }


def ensure_table_and_insert(
    source_file: str, payload: dict[str, Any], db_config: DbConfig
) -> str | None:
    sql_schema = "CREATE SCHEMA IF NOT EXISTS raw;"
    sql_extension = "CREATE EXTENSION IF NOT EXISTS pgcrypto;"
    sql_table = """
    CREATE TABLE IF NOT EXISTS raw.rating_assessments_history (
        id UUID PRIMARY KEY DEFAULT gen_random_uuid(),
        record_hash TEXT NOT NULL UNIQUE,
        entity_information TEXT NOT NULL,
        methodology TEXT NOT NULL,
        industry_risk TEXT NOT NULL,
        business_risk_profile TEXT NOT NULL,
        financial_risk_profile TEXT NOT NULL,
        credit_metrics TEXT NOT NULL,
        entity_name TEXT NOT NULL,
        country TEXT,
        industry TEXT,
        business_risk_score TEXT,
        financial_risk_score TEXT,
        rating_date DATE,
        source_system TEXT,
        document_version INTEGER NOT NULL DEFAULT 1 CHECK (document_version >= 1),
        ingested_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
        UNIQUE (entity_name, rating_date, document_version)
    );
    """
    sql_insert = """
    INSERT INTO raw.rating_assessments_history (
        record_hash,
        entity_information,
        methodology,
        industry_risk,
        business_risk_profile,
        financial_risk_profile,
        credit_metrics,
        entity_name,
        country,
        industry,
        business_risk_score,
        financial_risk_score,
        rating_date,
        source_system,
        document_version
    )
    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
    ON CONFLICT (record_hash) DO NOTHING
    RETURNING id;
    """

    conn_str = (
        f"host={db_config.host} "
        f"port={db_config.port} "
        f"dbname={db_config.dbname} "
        f"user={db_config.user} "
        f"password={db_config.password}"
    )

    with psycopg2.connect(conn_str) as conn:
        with conn.cursor() as cur:
            cur.execute(sql_schema)
            cur.execute(sql_extension)
            cur.execute(sql_table)
            entity_information = payload.get("entity_information", {})
            methodology = payload.get("methodology", {})
            industry_risk = payload.get("industry_risk", {})
            business_risk_profile = payload.get("business_risk_profile", {})
            financial_risk_profile = payload.get("financial_risk_profile", {})
            credit_metrics = payload.get("credit_metrics", [])
            canonical_payload = json.dumps(payload, sort_keys=True, separators=(",", ":"), ensure_ascii=True)
            record_hash = hashlib.sha256(canonical_payload.encode("utf-8")).hexdigest()

            cur.execute(
                sql_insert,
                (
                    record_hash,
                    json.dumps(entity_information, ensure_ascii=True),
                    json.dumps(methodology, ensure_ascii=True),
                    json.dumps(industry_risk, ensure_ascii=True),
                    json.dumps(business_risk_profile, ensure_ascii=True),
                    json.dumps(financial_risk_profile, ensure_ascii=True),
                    json.dumps(credit_metrics, ensure_ascii=True),
                    entity_information.get("name"),
                    entity_information.get("country_of_origin"),
                    entity_information.get("industry"),
                    business_risk_profile.get("overall_score"),
                    financial_risk_profile.get("overall_score"),
                    None,  # rating_date not present in source workbook
                    source_file,
                    1,
                ),
            )
            row = cur.fetchone()
            inserted_id = str(row[0]) if row else None
        conn.commit()

    return inserted_id


def main() -> None:
    data_dir = DATA_DIR.resolve()
    if not data_dir.exists() or not data_dir.is_dir():
        raise FileNotFoundError(f"Data directory not found: {data_dir}")

    workbook_paths = sorted(
        p
        for p in data_dir.iterdir()
        if p.is_file()
        and p.suffix.lower() in EXCEL_EXTENSIONS
        and not p.name.startswith("~$")
    )
    if not workbook_paths:
        raise FileNotFoundError(f"No Excel files found in {data_dir}")

    db_config = load_db_config(ENV_FILE.resolve())

    for workbook_path in workbook_paths:
        try:
            payload = extract_workbook(workbook_path, data_only=DATA_ONLY)
        except BadZipFile:
            print(f"Skipped invalid Excel file (not a zip workbook): {workbook_path}")
            continue

        inserted_id = ensure_table_and_insert(str(workbook_path), payload, db_config)
        if inserted_id is None:
            print(f"Skipped duplicate (same record_hash) for {workbook_path.name}")
        else:
            print(
                f"Inserted row id={inserted_id} into raw.rating_assessments_history for {workbook_path.name}"
            )


if __name__ == "__main__":
    main()
