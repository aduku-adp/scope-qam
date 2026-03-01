"""Excel extraction logic for MASTER sheet payload generation."""

from __future__ import annotations

import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


class MasterSheetExtractor:
    """Extract and normalize corporate rating data from the workbook MASTER sheet."""

    def __init__(self, target_sheet_name: str = "MASTER", data_only: bool = True) -> None:
        """Instantiate the extractor with workbook-read options."""
        self.target_sheet_name = target_sheet_name
        self.data_only = data_only

    @staticmethod
    def _parse_float_value(value: Any) -> float | str | None:
        """Parse numeric-like values, preserving non-numeric text when needed."""
        if value is None:
            return None
        if isinstance(value, str):
            text = value.strip()
            if not text or text.lower() in {"no data", "n/a", "na"}:
                return None
            try:
                return round(float(text.replace(",", "")), 3)
            except ValueError:
                return text
        if isinstance(value, (int, float)):
            return round(float(value), 3)
        return None

    @staticmethod
    def _parse_notches(value: Any) -> int | str | None:
        """Extract the first signed integer from a liquidity/notch cell."""
        if value is None:
            return None
        match = re.search(r"-?\d+", str(value))
        if not match:
            return str(value)
        return int(match.group(0))

    @staticmethod
    def _normalize_metric_name(metric_name: str) -> str:
        """Normalize free-text metric labels into snake_case identifiers."""
        normalized = metric_name.strip().lower()
        normalized = re.sub(r"[^a-z0-9]+", "_", normalized)
        return normalized.strip("_")

    @staticmethod
    def _get_label_value(master_sheet, label: str) -> Any:
        """Return the value from column 3 for a given label found in column 2."""
        for row_idx in range(1, master_sheet.max_row + 1):
            if master_sheet.cell(row=row_idx, column=2).value == label:
                return master_sheet.cell(row=row_idx, column=3).value
        return None

    def _extract_credit_metrics(self, master_sheet) -> list[dict[str, Any]]:
        """Extract credit metrics grid into normalized metric/value payload rows."""
        header_row = None
        for row_idx in range(1, master_sheet.max_row + 1):
            if master_sheet.cell(row=row_idx, column=2).value == "[Scope Credit Metrics]":
                header_row = row_idx
                break
        if header_row is None:
            return []

        years: list[str] = []
        col_idx = 3
        while True:
            year_value = master_sheet.cell(row=header_row, column=col_idx).value
            if year_value in (None, ""):
                break
            years.append(str(year_value))
            col_idx += 1

        # The lock flag is stored in the first column after the year value columns.
        lock_col = 3 + len(years)
        metrics: list[dict[str, Any]] = []

        row_idx = header_row + 1
        while row_idx <= master_sheet.max_row:
            metric_label = master_sheet.cell(row=row_idx, column=2).value
            if metric_label in (None, ""):
                break

            values = []
            for offset, year in enumerate(years):
                raw_value = master_sheet.cell(row=row_idx, column=3 + offset).value
                values.append({"year": year, "value": self._parse_float_value(raw_value)})

            lock_value = master_sheet.cell(row=row_idx, column=lock_col).value
            locked = str(lock_value).strip().lower() == "locked"

            metrics.append(
                {
                    "metric": self._normalize_metric_name(str(metric_label)),
                    "values": values,
                    "locked": locked,
                }
            )
            row_idx += 1

        return metrics

    def _extract_industry_risk(self, master_sheet) -> list[dict[str, Any]]:
        """Extract industry risk rows as a list of weighted industry risk items."""
        risk_row = None
        score_row = None
        weight_row = None

        for row_idx in range(1, master_sheet.max_row + 1):
            label = master_sheet.cell(row=row_idx, column=2).value
            if label == "Industry risk":
                risk_row = row_idx
            elif label == "Industry risk score":
                score_row = row_idx
            elif label == "Industry weight":
                weight_row = row_idx

        if risk_row is None or score_row is None or weight_row is None:
            return []

        items: list[dict[str, Any]] = []
        for col_idx in range(3, master_sheet.max_column + 1):
            classification = master_sheet.cell(row=risk_row, column=col_idx).value
            score = master_sheet.cell(row=score_row, column=col_idx).value
            weight = master_sheet.cell(row=weight_row, column=col_idx).value

            if classification in (None, "") and score in (None, "") and weight in (None, ""):
                if items:
                    break
                continue

            items.append(
                {
                    "industry_classification": None if classification in (None, "") else str(classification),
                    "industry_risk_score": None if score in (None, "") else str(score),
                    "industry_weight": self._parse_float_value(weight),
                }
            )

        return items

    def extract_workbook(self, workbook_path: Path) -> dict[str, Any]:
        """Extract a complete structured payload from one workbook."""
        workbook = load_workbook(workbook_path, data_only=self.data_only)
        if self.target_sheet_name not in workbook.sheetnames:
            raise ValueError(f"Sheet '{self.target_sheet_name}' not found in workbook")

        ws = workbook[self.target_sheet_name]

        methodologies: list[str] = []
        for col_idx in range(3, ws.max_column + 1):
            value = ws.cell(row=5, column=col_idx).value
            if value in (None, ""):
                continue
            methodologies.append(str(value))

        corporate_sector_value = ws.cell(row=3, column=3).value

        return {
            "company_information": {
                "name": self._get_label_value(ws, "Rated entity"),
                "corporate_sector": corporate_sector_value,
                "country_of_origin": self._get_label_value(ws, "Country of origin"),
                "reporting_currency": self._get_label_value(ws, "Reporting Currency/Units"),
                "accounting_principles": self._get_label_value(ws, "Accounting principles"),
                "fiscal_year_end": self._get_label_value(ws, "End of business year"),
                "methodology": {
                    "rating_methodologies_applied": methodologies,
                },
                "industry_risk": self._extract_industry_risk(ws),
                "segmentation_criteria": self._get_label_value(ws, "Segmentation criteria"),
            },
            "business_risk_profile": {
                "overall_score": self._get_label_value(ws, "Business risk profile"),
                "components": {
                    "blended_industry_risk_profile": self._get_label_value(
                        ws, "(Blended) Industry risk profile"
                    ),
                    "competitive_positioning": self._get_label_value(
                        ws, "Competitive Positioning"
                    ),
                    "market_share": self._get_label_value(ws, "Market share"),
                    "diversification": self._get_label_value(ws, "Diversification"),
                    "operating_profitability": self._get_label_value(
                        ws, "Operating profitability"
                    ),
                    "sector_company_specific_factors_1": self._get_label_value(
                        ws, "Sector/company-specific factors (1)"
                    ),
                    "sector_company_specific_factors_2": self._get_label_value(
                        ws, "Sector/company-specific factors (2)"
                    ),
                },
            },
            "financial_risk_profile": {
                "overall_score": self._get_label_value(ws, "Financial risk profile"),
                "components": {
                    "leverage": self._get_label_value(ws, "Leverage"),
                    "interest_cover": self._get_label_value(ws, "Interest cover"),
                    "cash_flow_cover": self._get_label_value(ws, "Cash flow cover"),
                    "liquidity_adjustment_notches": self._parse_notches(
                        self._get_label_value(ws, "Liquidity")
                    ),
                },
            },
            "credit_metrics": self._extract_credit_metrics(ws),
        }
